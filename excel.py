from __future__ import annotations

from pathlib import Path
from typing import Iterable

from clientes import formatar_documento, formatar_telefone
from database import connect


def _openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "A biblioteca openpyxl nao esta instalada. Instale com: pip install openpyxl"
        ) from exc

    return Workbook, Alignment, Font, PatternFill, get_column_letter


def _ajustar_planilha(ws, headers: Iterable[str], get_column_letter, Font, PatternFill, Alignment) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for index, header in enumerate(headers, start=1):
        coluna = get_column_letter(index)
        tamanho = max(
            len(str(ws.cell(row=row, column=index).value or ""))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[coluna].width = max(len(header) + 2, min(tamanho + 2, 45))


def exportar_clientes(pasta_destino: str | Path) -> Path:
    Workbook, Alignment, Font, PatternFill, get_column_letter = _openpyxl()
    destino = Path(pasta_destino) / "clientes.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)

    with connect() as connection:
        clientes = connection.execute(
            """
            SELECT id, nome, telefone, endereco, cpf_cnpj, data_cadastro
            FROM clientes
            ORDER BY nome COLLATE NOCASE
            """
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["ID", "Nome", "Telefone", "Endereco", "CPF/CNPJ", "Data de cadastro"]
    ws.append(headers)

    for cliente in clientes:
        ws.append(
            [
                cliente["id"],
                cliente["nome"],
                formatar_telefone(cliente["telefone"]),
                cliente["endereco"] or "",
                formatar_documento(cliente["cpf_cnpj"]),
                cliente["data_cadastro"] or "",
            ]
        )

    _ajustar_planilha(ws, headers, get_column_letter, Font, PatternFill, Alignment)
    wb.save(destino)
    return destino


def exportar_pedidos(pasta_destino: str | Path) -> Path:
    Workbook, Alignment, Font, PatternFill, get_column_letter = _openpyxl()
    destino = Path(pasta_destino) / "pedidos.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)

    with connect() as connection:
        pedidos = connection.execute(
            """
            SELECT
                pedidos.id,
                clientes.nome AS cliente_nome,
                pedidos.numero_pedido,
                pedidos.caminho_pdf,
                pedidos.data_pedido
            FROM pedidos
            LEFT JOIN clientes ON clientes.id = pedidos.cliente_id
            ORDER BY pedidos.id DESC
            """
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = ["ID", "Cliente", "Numero do pedido", "PDF", "Data do pedido"]
    ws.append(headers)

    for pedido in pedidos:
        ws.append(
            [
                pedido["id"],
                pedido["cliente_nome"] or "",
                pedido["numero_pedido"] or "",
                pedido["caminho_pdf"] or "",
                pedido["data_pedido"] or "",
            ]
        )

    _ajustar_planilha(ws, headers, get_column_letter, Font, PatternFill, Alignment)
    wb.save(destino)
    return destino
